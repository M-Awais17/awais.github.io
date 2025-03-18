from flask import Flask, render_template, request, redirect, url_for, session, flash
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with a secure random key

# File path for saving user data
user_data_file = 'data/users.xlsx'

# Function to initialize the Excel file if it doesn't exist or is invalid
def initialize_excel_file():
    try:
        # Try to open the existing file
        if os.path.exists(user_data_file):
            load_workbook(user_data_file).close()
        else:
            # Create and set up the file with headers if it doesn’t exist
            wb = Workbook()
            ws = wb.active
            ws.append(['UserID', 'First Name', 'Surname', 'Username', 'Country Code', 'Phone Number', 'Email', 'Password'])
            wb.save(user_data_file)
    except Exception as e:
        # If an error occurs, recreate the file
        wb = Workbook()
        ws = wb.active
        ws.append(['UserID', 'First Name', 'Surname', 'Username', 'Country Code', 'Phone Number', 'Email', 'Password'])
        wb.save(user_data_file)
        print(f"Error encountered: {e}. File recreated.")

# Ensure data directory exists
os.makedirs('data', exist_ok=True)
initialize_excel_file()

# Utility to save user data
def save_user_data(first_name, surname, username, country_code, phone, email, password):
    initialize_excel_file()  # Ensure the file is correctly initialized
    
    # Load the workbook and add user data
    wb = load_workbook(user_data_file)
    ws = wb.active
    user_id = ws.max_row  # New UserID based on row count
    new_user = [user_id, first_name, surname, username, country_code, phone, email, password]
    ws.append(new_user)
    wb.save(user_data_file)

# Utility to load user data and check credentials
def load_user_data():
    users = {}
    try:
        wb = load_workbook(user_data_file)
        ws = wb.active
        # Skip the header row and read data
        for row in ws.iter_rows(min_row=2, values_only=True):
            username, email, password = row[3], row[6], row[7]
            users[username] = {
                'email': email,
                'password': password
            }
    except Exception as e:
        flash("Error loading user data.")
        print("Error:", e)
    return users

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Retrieve form data
        first_name = request.form['name']
        surname = request.form['surname']
        username = request.form['username']
        country_code = request.form['country_code']
        phone = request.form['phone']
        email = request.form['email']
        password = request.form['password']
        
        # Check if user already exists
        users = load_user_data()
        if username in users:
            flash('Username already exists, please choose another one.')
            return redirect(url_for('register'))
        
        # Save new user data
        save_user_data(first_name, surname, username, country_code, phone, email, password)
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username_or_email = request.form['username_or_email']
        password = request.form['password']
        
        # Load user data to validate login
        users = load_user_data()
        
        # Check if the input is a username or an email and validate
        user_found = None
        for username, data in users.items():
            if username == username_or_email or data['email'] == username_or_email:
                user_found = data
                break

        if user_found and user_found['password'] == password:
            session['username'] = username  # Store the username in the session
            flash('Login successful!')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username/email or password.')
            return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        flash('You must be logged in to view this page.')
        return redirect(url_for('login'))
    
    return render_template('dashboard.html', username=session['username'])

@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('Logged out successfully.')
    return redirect(url_for('home'))

if __name__ == '__main__':
     app.run(debug=True, host="0.0.0.0", port=5000, 
        ssl_context=('/data/data/com.termux/files/home/.termux/ssl/server.crt', 
                      '/data/data/com.termux/files/home/.termux/ssl/server.key'))
